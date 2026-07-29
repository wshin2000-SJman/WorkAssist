import openpyxl
import os

folder = r"C:\Users\user\Desktop\Work\SJautomation\Antigravity\SJ_WorkAssist\workassist-v2"
master_file = "(삼정_MS사업부) 2026년 수주매출 현황_260518.xlsm"
path = os.path.join(folder, master_file)

# load without data_only=True to see formulas
wb = openpyxl.load_workbook(path, data_only=False)
sheet = wb["차트"]

print("=== Formula Inspection inside '차트' ===")
for r in range(13, 20):
    for c in range(1, 15):
        cell = sheet.cell(row=r, column=c)
        if cell.value is not None:
            print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: {cell.value!r}")

wb.close()
