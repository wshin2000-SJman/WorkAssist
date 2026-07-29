import openpyxl
import os

folder = r"C:\Users\user\Desktop\Work\SJautomation\Antigravity\SJ_WorkAssist\workassist-v2"
master_file = "(삼정_MS사업부) 2026년 수주매출 현황_260518.xlsm"
path = os.path.join(folder, master_file)

wb = openpyxl.load_workbook(path, data_only=False)
sheet = wb["차트"]

print("=== Cell Array Formulas ===")
for r in [15, 16]:
    for c in [3, 4]: # C and D
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        # Check if it's an ArrayFormula or has a text property
        if val is not None:
            if hasattr(val, "text"):
                print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: ArrayFormula: {val.text}")
            else:
                print(f"Cell {openpyxl.utils.get_column_letter(c)}{r}: {val!r}")

wb.close()
