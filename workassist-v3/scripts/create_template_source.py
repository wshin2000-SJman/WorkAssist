import openpyxl
import os

folder = r"C:\Users\user\Desktop\Work\SJautomation\Antigravity\SJ_WorkAssist\workassist-v2"
template_path = os.path.join(folder, "scripts", "order_template.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "수주매출 현황"

# Write Row 2 title
ws.cell(row=2, column=2, value="(MS사업부) 2026년 수주/ 매출 현황")

# Write Row 4 headers (21 columns)
headers = [
    "열1", "수주일자", "프로젝트코드", "구분", "고객사", 
    "고객 발주번호", "제품명", "고객 담당자", "영업담당자", "납품요청일", 
    "매입처", "수량", "수주단가", "수주금액", "납품 예정일", 
    "실제 납품일", "1차 계산서", "2차 계산서", "1차 정산금", "2차 정산금", "비고"
]

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=h)
    cell.font = openpyxl.styles.Font(bold=True)
    # Give it a nice clean border and background to look premium
    cell.fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Set columns width to look premium
for col_idx in range(1, len(headers) + 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15

wb.save(template_path)
print(f"Blank template file created at {template_path}")
wb.close()
