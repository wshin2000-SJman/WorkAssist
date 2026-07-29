import openpyxl
import os
import sys

folder = r"C:\Users\user\Desktop\Work\SJautomation\Antigravity\SJ_WorkAssist\workassist-v2"
master_file = "(삼정_MS사업부) 2026년 수주매출 현황_260518.xlsm"
path = os.path.join(folder, master_file)

wb = openpyxl.load_workbook(path, data_only=True)
# Find sheet name, usually it is "입력표"
sheet = wb["입력표"]

# Let's inspect the first 10 rows and find the header row (typically row 4) and first data rows
out_lines = []
out_lines.append(f"Sheet Name: {sheet.title}")
out_lines.append(f"Max Row: {sheet.max_row}")
out_lines.append(f"Max Column: {sheet.max_column}")

# Let's list the values of Row 4 (which contains headers in Korean)
headers = [sheet.cell(row=4, column=c).value for c in range(1, 35)]
out_lines.append("\n=== Headers (Row 4, Column 1 to 34) ===")
for idx, h in enumerate(headers):
    out_lines.append(f"Col {idx+1} ({openpyxl.utils.get_column_letter(idx+1)}): {h}")

out_lines.append("\n=== Data Row 5 ===")
for c in range(1, 35):
    val = sheet.cell(row=5, column=c).value
    out_lines.append(f"Col {c} ({headers[c-1]}): {val} (Type: {type(val).__name__})")

out_lines.append("\n=== Data Row 7 ===")
for c in range(1, 35):
    val = sheet.cell(row=7, column=c).value
    out_lines.append(f"Col {c} ({headers[c-1]}): {val} (Type: {type(val).__name__})")

output_path = os.path.join(folder, "scripts", "excel_headers_inspection.txt")
with open(output_path, "w", encoding="utf-8") as f:
    f.write("\n".join(out_lines))

print(f"Inspection complete. Written to {output_path}")
wb.close()
