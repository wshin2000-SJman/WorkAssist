import openpyxl
import os

folder = r"C:\Users\user\Desktop\Work\SJautomation\Antigravity\SJ_WorkAssist\workassist-v2"
master_file = "(삼정_MS사업부) 2026년 수주매출 현황_260518.xlsm"
path = os.path.join(folder, master_file)

wb = openpyxl.load_workbook(path, data_only=True)
print(f"Workbook sheets: {wb.sheetnames}")

out_lines = []

# Inspect '수주매출 현황'
if "수주매출 현황" in wb.sheetnames:
    sheet = wb["수주매출 현황"]
    out_lines.append(f"=== Sheet: {sheet.title} (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ===")
    
    # Read headers (usually rows 1 to 4 might have headers)
    for r in range(1, 6):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 25)]
        out_lines.append(f"Row {r}: {row_vals}")
        
    out_lines.append("\n=== Data Rows Sample ===")
    for r in range(6, 12):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 25)]
        out_lines.append(f"Row {r}: {row_vals}")

# Inspect '차트' or sheets starting with chart
chart_sheets = [s for s in wb.sheetnames if "차트" in s or "Chart" in s or s == "차트"]
for cs in chart_sheets:
    sheet = wb[cs]
    out_lines.append(f"\n=== Sheet: {sheet.title} (Rows: {sheet.max_row}, Cols: {sheet.max_column}) ===")
    for r in range(1, min(20, sheet.max_row + 1)):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(15, sheet.max_column + 1))]
        if any(v is not None for v in row_vals):
            out_lines.append(f"Row {r}: {row_vals}")

output_path = os.path.join(folder, "scripts", "sheets_details.txt")
with open(output_path, "w", encoding="utf-8") as f:
    f.write("\n".join(out_lines))

print(f"Inspection complete. Written to {output_path}")
wb.close()
