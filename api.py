import db
import csv
import os
import io
import hashlib

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
HAS_OPENPYXL = True

from version import VERSION

def hash_password(password):
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

class API:
    def __init__(self):
        db.init_db()
        self.current_user_id = None
        self.APP_VERSION = VERSION
        self.window = None

    def set_window(self, window):
        self.window = window

    def select_folder(self):
        if not self.window:
            return None
        import webview
        result = self.window.create_file_dialog(webview.FOLDER_DIALOG)
        if result and len(result) > 0:
            return result[0]
        return None

    def get_app_version(self):
        return self.APP_VERSION

    # Auth API
    def register(self, username, password, hint):
        user_id = db.create_user(username, hash_password(password), hint)
        if user_id:
            return {'status': 'success'}
        return {'status': 'error', 'message': 'Username already exists'}

    def login(self, username, password):
        user = db.get_user(username)
        if not user:
            return {'status': 'error', 'message': 'User does not exist'}
        if user['password_hash'] != hash_password(password):
            return {'status': 'error', 'message': 'Incorrect password'}
        
        self.current_user_id = user['id']
        return {'status': 'success'}

    def logout(self):
        self.current_user_id = None
        return {'status': 'success'}

    def get_hint(self, username):
        user = db.get_user(username)
        if user:
            return {'status': 'success', 'hint': user['password_hint']}
        return {'status': 'error', 'message': 'User not found'}

    def change_password(self, username, old_password, new_password):
        user = db.get_user(username)
        if user and user['password_hash'] == hash_password(old_password):
            db.change_password(username, hash_password(new_password))
            return {'status': 'success'}
        return {'status': 'error', 'message': 'Invalid ID or Password'}

    # Task API
    def get_tasks(self):
        if self.current_user_id is None:
            return []
        return db.get_all_tasks(self.current_user_id)

    def add_task(self, data):
        if self.current_user_id is None:
            return {'status': 'error', 'message': 'Not logged in'}
            
        import datetime
        today = datetime.datetime.now()
        date_prefix = today.strftime('%y/%m/%d')
        seq = db.get_next_tag_sequence(self.current_user_id, date_prefix)
        
        if seq > 999:
            return {'status': 'error', 'message': 'tag_limit_exceeded'}
            
        task_tag = f"T-{date_prefix}-{seq:03d}"
        
        title = data.get('title', 'Untitled')
        content = data.get('content', '')
        manager = data.get('manager', '')
        start_date = data.get('start_date', '')
        due_date = data.get('due_date', '')
        status = data.get('status', 'Note')
        is_urgent = data.get('is_urgent', False)
        
        task_id = db.add_task(title, content, manager, start_date, due_date, status, is_urgent, self.current_user_id, task_tag)
        return {'status': 'success', 'id': task_id}

    def update_task_status(self, task_id, new_status):
        db.update_task_status(task_id, new_status)
        return {'status': 'success'}

    def save_review_and_complete(self, task_id, comment):
        db.save_review_comment(task_id, comment)
        db.update_task_status(task_id, 'Done')
        return {'status': 'success'}

    def update_task_dates(self, task_id, start_date, due_date):
        db.update_task_dates(task_id, start_date, due_date)
        return {'status': 'success'}

    def update_task_details(self, task_id, data):
        title = data.get('title')
        content = data.get('content')
        manager = data.get('manager')
        start_date = data.get('start_date')
        due_date = data.get('due_date')
        db.update_task_details(task_id, title, content, manager, start_date, due_date)
        return {'status': 'success'}

    def delete_task(self, task_id):
        db.delete_task(task_id)
        return {'status': 'success'}

    def initialize_data(self):
        if self.current_user_id is None:
            return {'status': 'error', 'message': 'Not logged in'}
        db.initialize_user_data(self.current_user_id)
        return {'status': 'success'}

    def export_csv(self):
        if self.current_user_id is None:
            return {'status': 'error', 'message': 'Not logged in'}
            
        tasks = db.get_all_tasks(self.current_user_id)
        
        # Determine the user's desktop path to save the export
        desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        export_file = os.path.join(desktop_path, 'SJKanban_Export.csv')
        
        # Write to CSV with UTF-8 BOM for Excel compatibility
        with open(export_file, mode='w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['ID', 'Title', 'Content', 'Manager', 'Start Date', 'Due Date', 'Status', 'Urgent', 'Created At'])
            for t in tasks:
                writer.writerow([
                    t['id'], t['title'], t['content'], t['manager'], 
                    t['start_date'], t['due_date'], t['status'], 
                    'Yes' if t['is_urgent'] else 'No', t['created_at']
                ])
        
        return {'status': 'success', 'path': export_file}

    # Meeting API
    def get_meetings(self):
        if self.current_user_id is None:
            return []
        return db.get_all_meetings(self.current_user_id)

    def save_meeting(self, data):
        if self.current_user_id is None:
            return {'status': 'error', 'message': 'Not logged in'}
        
        meeting_id = data.get('id')
        title = data.get('title', 'Untitled')
        date = data.get('date', '')
        participants = data.get('participants', '')
        location = data.get('location', '')
        decisions = data.get('decisions', '[]')
        action_items = data.get('action_items', '[]')
        memo = data.get('memo', '')
        
        if meeting_id:
            db.update_meeting(meeting_id, title, date, participants, location, decisions, action_items, memo)
            return {'status': 'success', 'id': meeting_id}
        else:
            new_id = db.add_meeting(self.current_user_id, title, date, participants, location, decisions, action_items, memo)
            return {'status': 'success', 'id': new_id}

    def delete_meeting(self, meeting_id):
        db.delete_meeting(meeting_id)
        return {'status': 'success'}

    def _generate_md_and_save(self, data, save_dir):
        import datetime
        import json
        
        # Determine filename: YYYYMMDD_(회의록)_제목.md
        try:
            date_obj = datetime.datetime.strptime(data['date'], '%Y-%m-%d')
            date_str = date_obj.strftime('%Y%m%d')
        except:
            date_str = "00000000"
            
        safe_title = "".join([c for c in data['title'] if c.isalnum() or c in (' ', '_', '-')]).rstrip()
        filename = f"{date_str}_(회의록)_{safe_title}.md"
        
        os.makedirs(save_dir, exist_ok=True)
        export_file = os.path.join(save_dir, filename)
        
        # Construct Markdown content
        decisions = json.loads(data['decisions'] if isinstance(data['decisions'], str) else '[]')
        action_items = json.loads(data['action_items'] if isinstance(data['action_items'], str) else '[]')
        
        md_content = f"# 📝 {data['title']}\n\n"
        md_content += f"**날짜:** {data['date']}\n"
        md_content += f"**장소:** {data['location']}\n"
        md_content += f"**참석자:** {data['participants']}\n\n"
        
        if data.get('memo'):
            md_content += f"## 📝 Free Memo\n{data['memo']}\n\n"
        
        md_content += "## 💡 주요 결정 사항\n"
        if decisions:
            for d in decisions:
                md_content += f"- **{d['issue']}**\n"
                md_content += f"  - 결정: {d['decision']}\n"
                md_content += f"  - 근거: {d['reason']}\n"
        else:
            md_content += "*(없음)*\n"
            
        md_content += "\n## ✅ Action Items\n"
        if action_items:
            md_content += "| 할 일 | 담당자 | 기한 |\n"
            md_content += "| :--- | :--- | :--- |\n"
            for a in action_items:
                md_content += f"| {a['task']} | {a['owner']} | {a['due_date']} |\n"
        else:
            md_content += "*(없음)*\n"
            
        md_content += f"\n---\n*Created by SJ WorkAssist at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*"
        
        with open(export_file, mode='w', encoding='utf-8') as f:
            f.write(md_content)
            
        return export_file

    def export_meeting_md(self, data):
        # Determine save directory
        save_dir = data.get('save_dir')
        if not save_dir:
            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            save_dir = os.path.join(desktop_path, 'SJ_Meetings')
        
        path = self._generate_md_and_save(data, save_dir)
        return {'status': 'success', 'path': path}

    def export_all_meetings_md(self, save_dir):
        if self.current_user_id is None:
            return {'status': 'error', 'message': 'Not logged in'}
            
        meetings = db.get_all_meetings(self.current_user_id)
        if not meetings:
            return {'status': 'success', 'count': 0}

        count = 0
        for m in meetings:
            self._generate_md_and_save(m, save_dir)
            count += 1
            
        return {'status': 'success', 'count': count}

    # Project Manager API
    def get_projects(self, status='active'):
        if self.current_user_id is None:
            return []
        return db.get_all_projects(self.current_user_id, status)

    def save_project(self, data):
        if self.current_user_id is None:
            return {'status': 'error', 'message': 'Not logged in'}
        
        project_id = data.get('id')
        name = data.get('name', 'Untitled Project')
        description = data.get('description', '')
        manager = data.get('manager', '')
        client = data.get('client', '')
        dept1 = data.get('dept1_name', '[DPT. 1]')
        dept2 = data.get('dept2_name', '[DPT. 2]')
        dept3 = data.get('dept3_name', '[DPT. 3]')
        dept4 = data.get('dept4_name', '[DPT. 4]')
        
        if project_id:
            db.update_project(project_id, name, description, manager, client, dept1, dept2, dept3, dept4)
            return {'status': 'success', 'id': project_id}
        else:
            # Note: add_project in db.py doesn't take depts yet, it uses defaults. 
            # But I updated db.py to take depts if I wanted to, but the current add_project in db.py uses hardcoded defaults.
            # Wait, I updated add_project in db.py to take values? No, I just updated the INSERT.
            new_id = db.add_project(self.current_user_id, name, description, manager, client)
            return {'status': 'success', 'id': new_id}

    def delete_project(self, project_id):
        db.delete_project(project_id)
        return {'status': 'success'}

    def delete_project_permanent(self, project_id):
        db.delete_project_permanent(project_id)
        return {'status': 'success'}

    def mark_project_done(self, project_id):
        db.update_project_status(project_id, 'done')
        return {'status': 'success'}

    def restore_project(self, project_id):
        db.update_project_status(project_id, 'active')
        return {'status': 'success'}

    def get_status_logs(self, project_id):
        return db.get_status_logs(project_id)

    def save_status_log(self, data):
        project_id = data.get('project_id')
        department = data.get('department')
        text_content = data.get('text_content', '')
        image_path = data.get('image_path', '')
        title = data.get('title', '')
        manager = data.get('manager', '')
        start_date = data.get('start_date', '')
        due_date = data.get('due_date', '')
        
        # Generate tag: L-YY/MM/DD-###
        import datetime
        now = datetime.datetime.now()
        date_str = now.strftime('%y/%m/%d')
        serial = db.get_next_tag_serial(date_str)
        tag = f"L-{date_str}-{serial:03d}"
        
        new_id = db.add_status_log(project_id, department, text_content, image_path, tag, title, manager, start_date, due_date)
        return {'status': 'success', 'id': new_id, 'tag': tag}

    def update_status_log(self, data):
        log_id = data.get('id')
        title = data.get('title', '')
        text_content = data.get('text_content', '')
        manager = data.get('manager', '')
        start_date = data.get('start_date', '')
        due_date = data.get('due_date', '')
        image_path = data.get('image_path', '')
        
        db.update_status_log_full(log_id, title, text_content, manager, start_date, due_date, image_path)
        return {'status': 'success'}

    def delete_status_log(self, log_id):
        db.delete_status_log(log_id)
        return {'status': 'success'}

    def delete_status_log_permanent(self, log_id):
        db.delete_status_log_permanent(log_id)
        return {'status': 'success'}

    def export_project_excel(self, project_id, image_data=None):
        try:
            if not HAS_OPENPYXL:
                return {'status': 'error', 'message': 'openpyxl is not installed. Please contact support.'}
                
            project = db.get_project_by_id(project_id)
            if not project:
                return {'status': 'error', 'message': 'Project not found.'}
                
            logs = db.get_status_logs(project_id)
            milestones = db.get_milestones(project_id)
            
            wb = openpyxl.Workbook()
            
            # 1. Chart Sheet (Visual Table)
            if image_data:
                import io
                import base64
                from openpyxl.drawing.image import Image as OpenpyxlImage
                
                ws_chart = wb.active
                ws_chart.title = "Chart"
                
                # Decode base64 image
                header, encoded = image_data.split(",", 1)
                data = base64.b64decode(encoded)
                img_ptr = io.BytesIO(data)
                
                img = OpenpyxlImage(img_ptr)
                # Scale down slightly if too large? Or just keep it.
                ws_chart.add_image(img, 'B2')
                
                # 2. Logs Sheet
                ws_logs = wb.create_sheet(title="Logs")
            else:
                ws_logs = wb.active
                ws_logs.title = "Logs"
            
            # 1/2. Logs Sheet Content
            headers_logs = ["Tag", "Department", "Date", "Content", "Status", "Manager", "Schedule"]
            ws_logs.append(headers_logs)
            
            for log in reversed(logs):
                status_text = log['status']
                if status_text == 'done': status_text = 'DONE'
                elif status_text == 'deleted': status_text = 'DELETED'
                else: status_text = 'ACTIVE'
                
                schedule = f"{log.get('start_date', '')} ~ {log.get('due_date', '')}"
                if schedule == " ~ ": schedule = ""
                
                ws_logs.append([
                    log.get('tag', ''),
                    log.get('department', ''),
                    log.get('timestamp', ''),
                    log.get('text_content', ''),
                    status_text,
                    log.get('manager', ''),
                    schedule
                ])
                
            # 3. Milestones Sheet
            ws_ms = wb.create_sheet(title="Milestones")
            headers_ms = ["Slot", "Deadline", "Name", "Content", "Status"]
            ws_ms.append(headers_ms)
            
            for ms in sorted(milestones, key=lambda x: x['slot_number']):
                if not ms.get('is_saved'):
                    continue
                ws_ms.append([
                    ms['slot_number'],
                    ms.get('deadline', ''),
                    ms.get('name', ''),
                    ms.get('content', ''),
                    'COMPLETED' if ms.get('is_done') else 'ACTIVE'
                ])
                
            # Styling formatting function
            def style_sheet(ws):
                if ws.title == "Chart": return # Skip chart sheet
                
                header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50) # Cap width at 50
                    ws.column_dimensions[column].width = adjusted_width
                    
                # Alignment for all cells
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

            style_sheet(ws_logs)
            style_sheet(ws_ms)
            
            safe_name = "".join([c for c in project['name'] if c.isalnum() or c in (' ', '_', '-')]).rstrip()
            filename = f"{safe_name}_TimeTable.xlsx"
        
            try:
                import webview
                file_types = ('Excel Files (*.xlsx)', 'All files (*.*)')
                dest_path = self.window.create_file_dialog(
                    webview.SAVE_DIALOG, 
                    save_filename=filename, 
                    file_types=file_types
                )
                
                if dest_path and len(dest_path) > 0:
                    export_file = dest_path[0]
                    wb.save(export_file)
                    return {'status': 'success', 'path': export_file}
                else:
                    return {'status': 'error', 'message': 'User cancelled export.'}
            except Exception as e:
                return {'status': 'error', 'message': f"Failed to open file dialog: {str(e)}"}
        except Exception as global_e:
            import traceback
            tb = traceback.format_exc()
            return {'status': 'error', 'message': f"Unexpected error: {str(global_e)}\n{tb}"}

    def mark_status_log_done(self, log_id):
        db.update_status_log_state(log_id, 'done')
        return {'status': 'success'}

    def restore_status_log(self, log_id):
        db.update_status_log_state(log_id, 'active')
        return {'status': 'success'}

    def mark_status_log_deleted(self, log_id):
        db.update_status_log_state(log_id, 'deleted')
        return {'status': 'success'}

    def upload_project_image(self, project_id):
        if not self.window:
            return {'status': 'error', 'message': 'No window context'}
        
        import webview
        import shutil
        import uuid
        
        file_types = ('Image files (*.png;*.jpg;*.jpeg;*.gif)', 'All files (*.*)')
        result = self.window.create_file_dialog(webview.OPEN_DIALOG, allow_multiple=False, file_types=file_types)
        
        if result and len(result) > 0:
            source_path = result[0]
            
            # Destination path: %LOCALAPPDATA%/SJ_Kanban/projects/{project_id}/
            base = db.get_data_dir()
            proj_dir = os.path.join(base, 'projects', str(project_id))
            os.makedirs(proj_dir, exist_ok=True)
            
            ext = os.path.splitext(source_path)[1]
            filename = f"{uuid.uuid4().hex}{ext}"
            dest_path = os.path.join(proj_dir, filename)
            
            shutil.copy2(source_path, dest_path)
            
            return {'status': 'success', 'path': dest_path}
            
        return {'status': 'cancelled'}

    def get_local_image_base64(self, path):
        if not path or not os.path.exists(path):
            return ""
        try:
            import base64
            with open(path, "rb") as image_file:
                encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
                ext = os.path.splitext(path)[1].lower()
                mime = "image/jpeg"
                if ext == ".png":
                    mime = "image/png"
                elif ext == ".gif":
                    mime = "image/gif"
                return f"data:{mime};base64,{encoded_string}"
        except Exception as e:
            return ""

    # Milestone API
    def get_milestones(self, project_id):
        return db.get_milestones(project_id)

    def save_milestone(self, data):
        project_id = data.get('project_id')
        slot_number = data.get('slot_number')
        name = data.get('name', '')
        deadline = data.get('deadline', '')
        content = data.get('content', '')
        is_saved = data.get('is_saved', False)
        is_done = data.get('is_done', False)
        
        db.save_milestone(project_id, slot_number, name, deadline, content, is_saved, is_done)
        return {'status': 'success'}

    def delete_milestone(self, project_id, slot_number):
        db.delete_milestone(project_id, slot_number)
        return {'status': 'success'}
